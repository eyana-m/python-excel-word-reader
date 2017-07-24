import openpyxl
import os

#os.chdir("\\")

workbook = openpyxl.Workbook()

print workbook
print workbook.get_sheet_names()

sheet = workbook.get_sheet_by_name('Sheet')

print sheet['A1'].value == None

sheet['A1'] = "Sup"

sheet['A2'] = "Friend"

os.chdir("Result/")

workbook.save('example.xlsx')

sheet2 = workbook.create_sheet()

print workbook.get_sheet_names()

sheet2.title = "Eyana"

print workbook.get_sheet_names()

workbook.save('example2.xlsx')

workbook.create_sheet(index=0,title="Mallari")
workbook.save('example3.xlsx')